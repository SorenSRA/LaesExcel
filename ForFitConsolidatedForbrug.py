import os
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

output_wb_skab = 'LifeForFitForbrugSkabelon.xlsx'
output_wb = 'LifeForFitConsol2021Q4.xlsx'

output_valuta = 'Euro'  # Enten 'Euro' eller 'Dkr'
euro_til_dkr = 7.45

partner_liste = {
    'SLS': '1. SaltenLangsø',
    'EPA': '2. Miljøstyrelsen',
    'SHL': '3. SHLF',
}

dir_base = 'C:\\SRA\\LIFE-ForFit Financial Reporting\\'
dir_report = '\\1. Financial Report'
dkr = 'Dkr'


# Indlæs Excel-sheet til dataframe
def get_input_df(wb):
    cols_laes = 'A:E'
    cols_navne = ['A', 'B', 'C', 'D', 'E']
    rk_start_skip = 0
    rk_slut_skip = 0
    laes_sh = 'Individual Cost Statement'

    return pd.read_excel(wb,
                         sheet_name=laes_sh,
                         skiprows=rk_start_skip,
                         usecols=cols_laes,
                         skipfooter=rk_slut_skip,
                         header=0,
                         names=cols_navne)


def get_valuta(df):
    kol = 'E'
    rk = 36

    return df.loc[rk, kol]


def get_partner(df):
    kol = 'B'
    rk = 4

    return df.loc[rk, kol]


def get_periode_to(df):
    kol = 'E'
    rk = 2

    return df.loc[rk, kol]


def get_output_fane(wb, sh_name):
    output_skabelon_sh = 'StdSheet'
    if sh_name not in wb.sheetnames:
        wb.copy_worksheet(wb[output_skabelon_sh]).title = sh_name

    return wb[sh_name]


def kopier_celle(outputfane, tekst, celle):
    xy = coordinate_from_string(celle)  # returns ('A',4)
    col = column_index_from_string(xy[0])  # returns 1
    row = xy[1]  # returns 4
    outputfane.cell(row=row, column=col).value = tekst

    return


def kopier_df_excel(outputfane, input_dataframe, cost=True):
    if cost:
        rk = 11
        kol_just = 2
    else:
        rk = 11
        kol_just = 5
    for row_str in dataframe_to_rows(input_dataframe, index=False, header=False):
        for kol, cel in enumerate(row_str):
            outputfane.cell(row=rk, column=kol + kol_just).value = int(cel)
        rk += 1

    return


def get_partner_input_wb(partn):
    path = dir_base + partn + dir_report
    print('2 ' + key + str(os.listdir(path)))
    for file in os.listdir(path):
        if os.path.splitext(file)[1] == '.xlsx':
            excel_fil_navn = str(os.path.join(path, file))
            print('3' + '  ' + excel_fil_navn)
            return str(os.path.join(path, file))


def get_valuta_kor(input_val, output_val):
    if output_val == dkr:  # output_valuta er Dkr
        if input_val == dkr:  # input_valuta er Dkr
            val_kor = 1
        else:  # input_valuta er EURO
            val_kor = euro_til_dkr
    else:  # output_valuta er EURO
        if input_val == dkr:
            val_kor = 1.0 / euro_til_dkr
        else:
            val_kor = 1.0
    return val_kor


def get_cost(df, cost=True):
    if cost:
        col = 'B'
        df_cost = df.loc[9:21, [col]]
    else:
        col = 'E'
        df_cost = df.loc[9:12, [col]]
    input_valuta = get_valuta(df)
    valuta_kor = get_valuta_kor(input_valuta, 'EURO')
    print(f'Input valuta: {input_valuta}  Output valuta: EURO  Valutakorrektion: {valuta_kor}')
    talrk = []
    for i in df_cost[col]:
        talrk.append(round(i * valuta_kor))
    df_cost.insert(1, 'EURO', talrk)
    talrk = []
    for i in df_cost['EURO']:
        talrk.append(round(i * euro_til_dkr))
    df_cost.insert(2, 'Dkr', talrk)
    df_cost = df_cost.drop(columns=[col])
    return df_cost


# åbn Excell-fil med standardtekst/-tabel
wb_output = load_workbook(output_wb_skab)

# opret dataframe til løbende summation af de enkelte partneres indlæste dataframes
ialt_cost_df = pd.DataFrame()
ialt_income_df = pd.DataFrame()

for key, partner in partner_liste.items():
    # dan en partnerspecifik fane i destinations-excell-filen
    output_sh = get_output_fane(wb_output, key)

    input_wb = get_partner_input_wb(partner)
    input_df = get_input_df(input_wb)
    partner_navn = get_partner(input_df)
    print(f'Partner: {partner_navn}')
    input_cost = get_cost(input_df, True)
    input_income = get_cost(input_df, False)

    kopier_celle(output_sh, get_partner(input_df), 'B6')
    kopier_celle(output_sh, get_periode_to(input_df), 'E4')

    # kopier omkostninger/income Euro/Dkr til Partnerfanen
    kopier_df_excel(output_sh, input_cost, cost=True)
    kopier_df_excel(output_sh, input_income, cost=False)

    # løbende summering af input_dataframes for hver enkelt partner
    ialt_cost_df = ialt_cost_df.add(input_cost, fill_value=0)
    # løbende summering af input_dataframes for hver enkelt partner
    ialt_income_df = ialt_income_df.add(input_income, fill_value=0)


output_sh = get_output_fane(wb_output, 'I alt')

kopier_celle(output_sh, 'Alle Partnere', 'B6')

kopier_df_excel(output_sh, ialt_cost_df, cost=True)
kopier_df_excel(output_sh, ialt_income_df, cost=False)

# gem Excell-fil med de kopierede forbrugs-/budgettal
wb_output.save(output_wb)

#"C:\Program Files\ArcGIS\Pro\bin\Python\envs\arcgispro-py3\Python.exe" LIFEActionForbrug.py
import os
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

projekt = 'OpenWoods'

if projekt == 'ForFit':
    from ForFitPartneroversigt import partnerliste, pathbase, pathspecff, output_wb, output_wb_skab
elif projekt == 'OpenWoods':   
    from OpenWoodsPartneroversigt import partnerliste, pathbase, pathspecff, output_wb, output_wb_skab
else:
    print('Fejl')


output_valuta = 'Dkr'  # Enten 'Euro' eller 'Dkr'
euro_til_dkr = 7.4393

dkr = 'Dkr'

def get_input_df(wb):
    cols_laes = 'C:Z'
    rk_start_skip = 3
    rk_slut_skip = 2
    laes_sh = 'BudgetDisponering'

    return pd.read_excel(wb,
                         sheet_name=laes_sh,
                         skiprows=rk_start_skip,
                         usecols=cols_laes,
                         skipfooter=rk_slut_skip)


def get_valuta(wb):
    valuta_sh = 'Individual Cost Statement'
    euro_kol = 'E'
    euro_rk = 37
    valuta_df = pd.read_excel(wb,
                              sheet_name=valuta_sh,
                              usecols=euro_kol,
                              header=None,
                              names=[euro_kol])

    return valuta_df.loc[euro_rk, euro_kol]


def get_output_fane(wb, sh_name):
    output_skabelon_sh = 'BudgetDisponering'
    if sh_name not in wb.sheetnames:
        wb.copy_worksheet(wb[output_skabelon_sh]).title = sh_name

    return wb[sh_name]


def kopier_df_excel(outputfane, input_dataframe):
    rk = 5
    for row_str in dataframe_to_rows(input_dataframe, index=False, header=False):
        for kol, cel in enumerate(row_str):
            outputfane.cell(row=rk, column=kol + 3).value = int(cel)
        rk += 1

    return


def get_partner_input_wb(partn):
    path = pathbase + partn + pathspecff
    print('2 ' + key + str(os.listdir(path)))
    for file in os.listdir(path):
        if os.path.splitext(file)[1] == '.xlsx':
            excel_fil_navn = str(os.path.join(path, file))
            print('3' + '  ' + excel_fil_navn)
            return str(os.path.join(path, file))


def get_valuta_kor(input_val):
    if output_valuta == dkr:  # output_valuta er Dkr
        if input_val == dkr:  # input_valuta er Dkr
            val_kor = 1
        else:  # input_valuta er EURO
            val_kor = euro_til_dkr
    else:  # output_valuta er EURO
        if input_val == dkr:
            val_kor = 1.0/euro_til_dkr
        else:
            val_kor = 1.0
    return val_kor


# åbn Excell-fil med standardtekst/-tabel
wb_output = load_workbook(output_wb_skab)

# opret dataframe til løbende summation af de enkelte partneres indlæste dataframes
ialt_df = pd.DataFrame()

for key, partner in partnerliste.items():
    input_wb = get_partner_input_wb(partner)
    input_df = get_input_df(input_wb)
    input_valuta = get_valuta(input_wb)
    valuta_kor = get_valuta_kor(input_valuta)
    print(f'Input valuta: {input_valuta}  Output valuta: {output_valuta}  Valutakorrektion: {valuta_kor}')
    input_df = input_df * valuta_kor

    sh_outputfane = get_output_fane(wb_output, key)

    kopier_df_excel(sh_outputfane, input_df)

    # løbende summering af input_dataframes for hver enkelt partner
    ialt_df = ialt_df.add(input_df, fill_value=0)


sh_outputfane = get_output_fane(wb_output, 'I alt')
kopier_df_excel(sh_outputfane, ialt_df)

# gem Excell-fil med de kopierede forbrugs-/budgettal
wb_output.save(output_wb)
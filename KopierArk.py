import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import rows_from_range

dirbase = 'C:\\SRA\\LIFE-ForFit Financial Reporting\\'
dirreport = '\\1. Financial Report'
# outputexcel = 'C:\\SRA\Forbrug\\NyLifeForFitForbrug.xlsx'

partnerliste = {
    'SLS': '1. SaltenLangsø',
    'EPA': '2. Miljøstyrelsen',
}

outputexcel = 'LifeForFitForbrug.xlsx'

inputark = 'Forudsatninger'

inputark_sh = 'ActionForbrugTD'

range_str = "C6:K18"

def kopier_excl_data():
    for key, partner in partnerliste.items():
        cwd = dirbase + partner + dirreport
        print('2 ' + key + str(os.listdir(cwd)))

        for f in os.listdir(cwd):
            if os.path.splitext(f)[1] == '.xlsx':
                print('Dette er en EXCEL-fil')
                print('3 ' + os.path.splitext(f)[0])
                wb_output.create_sheet(key)
                wb_input = load_workbook(os.path.join(cwd, f), data_only=True)
                sh_inputfane = wb_input[inputark_sh]
                sh_outputfane = wb_output[key]
                kopi_range(range_str, sh_inputfane, sh_outputfane)
            else:
                print('Dette er ikke en excel-fil', f)
    return


def kopi_range(omraade, sh_input, sh_output):
    for row in rows_from_range(omraade):
        print(row)
        for cell in row:
            sh_output[cell].value = sh_input[cell].value
            print(cell)
    return


# Selve programmet


wb_output = Workbook()

kopier_excl_data()

wb_output.save('Test.xlsx')


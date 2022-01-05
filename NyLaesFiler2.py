import os
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

dirbase = 'C:\\SRA\\LIFE-ForFit Financial Reporting\\'
dirreport = '\\1. Financial Report'
wb_forbrug_excel_skabelon = 'C:\\SRA\Forbrug\\LifeForFitForbrugSkabelon.xlsx'

wb_forbrug_excel = 'C:\\SRA\Forbrug\\LifeForFitForbrug.xlsx'

partnerliste = {
    'SLS': '1. SaltenLangsø',
    'EPA': '2. Miljøstyrelsen',
    'SHL': '3. SHLF',
}
paradigme_sh = 'StdSheet'
consolidate_sh = 'StdConsol'
consolidate_sh_title = 'Consolideret'
ialt_sh_title = 'I alt'
inputark_sh = 'Individual Cost Statement'
ini_row = 4

infoceller = {
    'b6': 'b6',  # Beneficiary
    'b7': 'b7',  # Type of Beneficiary
    'b8': 'b8',  # Legal status
    'e4': 'e4',  # To
    'e38': 'e6',  # KursIark
    'e7': 'e7',  # Local currency
    'a38': 'g6',  # EuroKurs
}

celle_eurokurs = 'a38'
celle_kurs_i_ark = 'e38'

okonomicelle = {
    'b11': 'c4',  # Personnel Cost
    'b12': 'd4',  # Personnel Cost - Nonadditional
    'b13': 'e4',  # Personnel Cost - Additional
    'b14': 'f4',  # Travel
    'b15': 'g4',  # ExAss
    'b16': 'h4',  # Infra
    'b17': 'i4',  # Equip
    'b18': 'j4',  # Proto
    'b20': 'l4',  # Consume
    'b21': 'm4',  # Other
    'b22': 'n4',  # Overhead
    'e11': 'q4',  # EU-contribution
    'e12': 'r4',  # Own-contribution
    'e13': 's4',  # Co-finans-contribution
    'e14': 't4',  # Direct income
}


def kopier_excl_data():
    rk = ini_row
    forbrug.copy_worksheet(forbrug[consolidate_sh]).title = consolidate_sh_title
    sh_consolidate = forbrug[consolidate_sh_title]

    forbrug.copy_worksheet(forbrug[paradigme_sh]).title = ialt_sh_title
    sh_ialt = forbrug[ialt_sh_title]

    for key, partner in partnerliste.items():
        cwd = dirbase + partner + dirreport
        print('2 ' + key + str(os.listdir(cwd)))

        for f in os.listdir(cwd):
            if os.path.splitext(f)[1] == '.xlsx':
                print('3 ' + os.path.splitext(f)[0])
                print('Dette er en gyldig - xlsx - EXCEL-fil')
                wb = load_workbook(os.path.join(cwd, f), data_only=True)
                sh_inputfane = wb[inputark_sh]
                forbrug.copy_worksheet(forbrug[paradigme_sh]).title = key
                sh_outputfane = forbrug[key]

                for incelle, outcelle in infoceller.items():
                    sh_outputfane[outcelle] = sh_inputfane[incelle].value

                print('4. Infoafsnit kopieret - for: ', f)

                if sh_inputfane[celle_kurs_i_ark].value == "Dkr":
                    valua_korrektion = 1/sh_inputfane[celle_eurokurs].value
                else:
                    valua_korrektion = 1

                eurokurs = sh_inputfane[celle_eurokurs].value

                for incelle, outcelle in okonomicelle.items():
                    sh_outputfane[incelle] = sh_inputfane[incelle].value * valua_korrektion
                    if sh_ialt[incelle].value is None:
                        sh_ialt[incelle].value = sh_outputfane[incelle].value
                    else:
                        sh_ialt[incelle].value = sh_ialt[incelle].value + sh_outputfane[incelle].value
                    col_str, row = coordinate_from_string(incelle)
                    col = column_index_from_string(col_str)+1  # off_set = 1 kolonne til højre
                    sh_outputfane.cell(row=row, column=col).value = sh_inputfane[incelle].value * valua_korrektion * eurokurs
                    if sh_ialt.cell(row=row, column=col).value is None:
                        sh_ialt.cell(row=row, column=col).value = sh_outputfane.cell(row=row, column=col).value
                    else:
                        sh_ialt.cell(row=row, column=col).value = sh_ialt.cell(row=row, column=col).value + sh_outputfane.cell(row=row, column=col).value

                print('4. Økonomiafsnit kopieret for: ', f)

                # opdaterer consolideringsfanen
                for incelle, outcelle in okonomicelle.items():
                    col_str = coordinate_from_string(outcelle)[0]
                    col = column_index_from_string(col_str)
                    sh_consolidate.cell(row=rk, column=2).value = sh_outputfane['b6'].value
                    sh_consolidate.cell(row=rk, column=col).value = sh_outputfane[incelle].value

                rk += 1
            else:
                print('9. Dette er ikke en gyldig - .xlsx -  excel-fil', f)


# Selve programmet

forbrug = load_workbook(wb_forbrug_excel_skabelon)

kopier_excl_data()

forbrug.save(wb_forbrug_excel)
